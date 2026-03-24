"""
Trends — Import historical production data from Excel to JSON.
Reads the production spreadsheet and generates a JSON file ready for the API import endpoint.

Usage:
    python import_pedidos.py

Output:
    pedidos_import.json — JSON array of parsed pedido records
"""

import json
import re
import sys
from datetime import datetime, date
from pathlib import Path

import openpyxl
from openpyxl.chartsheet import Chartsheet

# ── Config ───────────────────────────────────────────────────────────────────

EXCEL_PATH = (
    Path(__file__).parent / "Reunião" / "TRENDS - TABELA PRODUCAO 2026 - LUIS.xlsx"
)
OUTPUT_PATH = Path(__file__).parent / "pedidos_import.json"

# Tabs to skip (not monthly production)
SKIP_TABS = {
    "INICIO",
    "FACA BATIDA",
    "FACAS",
    "PAPEL",
    "PRECOS",
    "PRECOS (2)",
    "Gráf1",
    "Gráf2",
    "Gráf3",
    "CAIXA DE PAPELA",
    "FORNECEDOR",
    "COMPRAS",
}

# Section header keywords → tipo_producao mapping
SECTION_KEYWORDS = {
    "GRAFICA": "GRAFICAS",
    "GRAFICAS": "GRAFICAS",
    "GRAFICAS TERCEIROS": "GRAFICAS",
    "GRAFICA 10": "GRAFICAS",
    "FLEXO": "FLEXO",
    "LASER": "LASER",
    "PERSONALIZAR": "PERSONALIZAR",
    "BERFLAY": "PERSONALIZAR",
    "LABEL": "LABEL",
    "SUPERFLEXO": "FLEXO",
}

# Month name mapping for mes_referencia parsing
MONTH_MAP = {
    "DEZEMBRO": 12,
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARCO": 3,
    "MARÇO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
}


def detect_format(ws):
    """Detect column format based on header row."""
    d1 = str(ws.cell(1, 4).value or "").strip().upper()
    if d1 == "OK":
        return 3  # 2025-2026 format: A=client, B=date, C=date, D=OK, E=O/F, ...
    elif d1 in ("O/F", "OC", "OF"):
        return 2  # 2020-2024 standard: A=client, B=date, C=date, D=O/F, E=material, ...
    # Check if it looks like a monthly production tab at all
    b1 = str(ws.cell(1, 2).value or "").strip()
    if "Data" in b1 or "Entr" in b1:
        return 2  # Default to format 2
    return 0  # Not a production tab


def get_column_map(fmt):
    """Return column index (1-based) mapping for each format."""
    if fmt == 3:
        return {
            "cliente": 1,
            "data_entrega": 2,
            "data_producao": 3,
            "ok": 4,
            "of": 5,
            "material": 6,
            "codigo": 7,
            "etiqueta": 8,
            "quantidade": 9,
            "produzido_por": 10,
            "oc": 11,
            "valor": 12,
        }
    else:  # Format 2 (and 1 with minor differences)
        return {
            "cliente": 1,
            "data_entrega": 2,
            "data_producao": 3,
            "of": 4,
            "material": 5,
            "codigo": 6,
            "etiqueta": 7,
            "quantidade": 8,
            "ok": 9,
            "produzido_por": 10,
            "oc": 11,
            "valor": 12,
        }


def parse_date(val):
    """Parse a date value from Excel (could be datetime, string, or None)."""
    if val is None:
        return None
    if isinstance(val, datetime):
        # Filter out bogus dates (like 1900-01-03 from Excel errors)
        if val.year < 2019 or val.year > 2027:
            return None
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        if val.year < 2019 or val.year > 2027:
            return None
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    # Try dd/mm or dd/m format (short dates like "14/12", "09/1")
    m = re.match(r"^(\d{1,2})/(\d{1,2})$", s)
    if m:
        return None  # Partial date, we'll fill from month context
    # Try full date formats
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
        try:
            dt = datetime.strptime(s, fmt)
            if 2019 <= dt.year <= 2027:
                return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def parse_partial_date(val, year, month):
    """Parse partial dates like '14/12' or '09/1' using context year/month."""
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return parse_date(val)
    s = str(val).strip()
    m = re.match(r"^(\d{1,2})/(\d{1,2})$", s)
    if m:
        day = int(m.group(1))
        mon = int(m.group(2))
        if 1 <= day <= 31 and 1 <= mon <= 12:
            try:
                dt = date(year, mon, day)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                pass
    return parse_date(val)


def parse_quantity(val):
    """Parse quantity — could be number, string with dots/commas, or text like '263,15'."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace(" ", "")
    if not s:
        return None
    # Handle Brazilian number format: "263,15" or "5.600" or "15.600"
    # If has comma, treat as decimal separator
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_valor(val):
    """Parse monetary value."""
    return parse_quantity(val)


def is_section_header(val):
    """Check if a cell value is a section header (tipo_producao)."""
    if val is None:
        return None
    s = str(val).strip().upper()
    # Direct match
    if s in SECTION_KEYWORDS:
        return SECTION_KEYWORDS[s]
    # Partial match
    for keyword, tipo in SECTION_KEYWORDS.items():
        if keyword in s:
            return tipo
    return None


def is_data_row(row_data, col_map):
    """Check if a row contains actual production data."""
    etiqueta = row_data.get("etiqueta")
    material = row_data.get("material")
    codigo = row_data.get("codigo")
    quantidade = row_data.get("quantidade")

    # Must have at least etiqueta or (material + something else)
    if etiqueta and str(etiqueta).strip():
        return True
    if material and str(material).strip() and (codigo or quantidade):
        return True
    return False


def parse_tab_name(name):
    """Extract year and month from tab name like 'MARÇO 26' or 'JANEIRO23'."""
    name = name.strip().upper()
    # Try "MONTH YY" or "MONTHYY"
    for month_name, month_num in MONTH_MAP.items():
        if name.startswith(month_name):
            rest = name[len(month_name) :].strip()
            if rest:
                try:
                    yr = int(rest)
                    if yr < 100:
                        yr += 2000
                    return yr, month_num
                except ValueError:
                    pass
    return None, None


def cell_val(ws, row, col):
    """Get cell value, return None for empty."""
    v = ws.cell(row, col).value
    if v is None:
        return None
    if isinstance(v, str) and not v.strip():
        return None
    return v


def parse_sheet(ws, tab_name):
    """Parse a monthly production sheet into a list of pedido records."""
    fmt = detect_format(ws)
    if fmt == 0:
        return []

    col_map = get_column_map(fmt)
    year, month = parse_tab_name(tab_name)
    if year is None:
        return []

    records = []
    current_section = None
    current_client = None
    seq = 0

    # Start from row 2 (skip header)
    for row_idx in range(2, ws.max_row + 1):
        # Read all relevant cells
        row_data = {}
        for field, col_idx in col_map.items():
            row_data[field] = cell_val(ws, row_idx, col_idx)

        # Check if this row is a section header
        a_val = cell_val(ws, row_idx, 1)
        section = is_section_header(a_val)
        if section is not None:
            # Check if this row ONLY has the section name (no other meaningful data)
            has_data = any(
                cell_val(ws, row_idx, col_map[f]) is not None
                for f in ["material", "etiqueta", "quantidade"]
            )
            if not has_data:
                current_section = section
                continue

        # Check if this is a data row
        if not is_data_row(row_data, col_map):
            continue

        # Client handling: inherit from above if empty
        cliente = (
            str(row_data.get("cliente") or "").strip()
            if row_data.get("cliente")
            else None
        )
        # Filter out numeric-only values in client column (like "2120", "650", "1300", "0")
        if cliente and re.match(r"^[\d.,\s]+$", cliente):
            cliente = None
        # Filter out section headers that slipped through
        if cliente and is_section_header(cliente):
            current_section = is_section_header(cliente)
            cliente = None
        if cliente:
            current_client = cliente
        else:
            cliente = current_client

        if not cliente:
            continue

        # Parse fields
        of_val = row_data.get("of")
        of_str = str(of_val).strip() if of_val is not None else None
        if of_str and of_str in ("0", "#REF!"):
            of_str = None

        material = str(row_data.get("material") or "").strip() or None
        codigo = (
            str(row_data.get("codigo") or "").strip()
            if row_data.get("codigo")
            else None
        )
        if codigo and codigo in ("0", "#REF!"):
            codigo = None

        etiqueta = str(row_data.get("etiqueta") or "").strip() or None
        quantidade = parse_quantity(row_data.get("quantidade"))
        valor = parse_valor(row_data.get("valor"))

        produzido_por = str(row_data.get("produzido_por") or "").strip() or None
        oc = str(row_data.get("oc") or "").strip() if row_data.get("oc") else None

        ok_val = str(row_data.get("ok") or "").strip().upper()
        is_ok = ok_val in ("OK",)

        data_entrega = parse_partial_date(row_data.get("data_entrega"), year, month)
        data_producao = parse_partial_date(row_data.get("data_producao"), year, month)

        # Determine status
        status = "entregue" if is_ok else "pendente"

        # Generate unique numero
        seq += 1
        mes_code = f"{year}{month:02d}"
        numero = f"HIST-{mes_code}-{seq:04d}"

        mes_referencia = tab_name.strip()

        record = {
            "numero": numero,
            "cliente_nome": cliente,
            "status": status,
            "valor_total": valor,
            "data_entrega": data_entrega,
            "data_producao": data_producao,
            "ordem_fabricacao": of_str,
            "material": material,
            "codigo_faca": codigo,
            "etiqueta_dimensao": etiqueta,
            "quantidade": quantidade,
            "produzido_por": produzido_por or current_section,
            "tipo_producao": current_section,
            "ordem_compra": oc,
            "mes_referencia": mes_referencia,
            "origem": "importacao",
        }
        records.append(record)

    return records


def main():
    print(f"Lendo planilha: {EXCEL_PATH}")
    if not EXCEL_PATH.exists():
        print(f"ERRO: Arquivo não encontrado: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(str(EXCEL_PATH), data_only=True)

    all_records = []
    tabs_processed = 0
    tabs_skipped = 0

    for tab_name in wb.sheetnames:
        if tab_name in SKIP_TABS:
            tabs_skipped += 1
            continue

        ws = wb[tab_name]
        if isinstance(ws, Chartsheet):
            tabs_skipped += 1
            continue

        year, month = parse_tab_name(tab_name)
        if year is None:
            tabs_skipped += 1
            continue

        records = parse_sheet(ws, tab_name)
        if records:
            tabs_processed += 1
            all_records.extend(records)
            print(f"  {tab_name}: {len(records)} pedidos")
        else:
            tabs_skipped += 1

    print(f"\n=== Resumo ===")
    print(f"  Abas processadas: {tabs_processed}")
    print(f"  Abas ignoradas:   {tabs_skipped}")
    print(f"  Total de pedidos: {len(all_records)}")

    # Stats
    clientes = set(r["cliente_nome"] for r in all_records if r["cliente_nome"])
    materiais = set(r["material"] for r in all_records if r["material"])
    tipos = set(r["tipo_producao"] for r in all_records if r["tipo_producao"])

    print(f"  Clientes únicos:  {len(clientes)}")
    print(f"  Materiais únicos: {len(materiais)}")
    print(f"  Tipos produção:   {tipos}")

    # Status distribution
    status_count = {}
    for r in all_records:
        s = r["status"]
        status_count[s] = status_count.get(s, 0) + 1
    print(f"  Status: {status_count}")

    # Year distribution
    year_count = {}
    for r in all_records:
        yr = r["mes_referencia"][:4] if r["mes_referencia"] else "?"
        # Extract year from numero
        parts = r["numero"].split("-")
        if len(parts) >= 2:
            yr = parts[1][:4]
        year_count[yr] = year_count.get(yr, 0) + 1
    print(f"  Por ano: {dict(sorted(year_count.items()))}")

    # Write JSON
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(all_records, f, ensure_ascii=False, indent=2)

    print(f"\nJSON gerado: {OUTPUT_PATH}")
    print(f"Tamanho: {OUTPUT_PATH.stat().st_size / 1024:.1f} KB")


if __name__ == "__main__":
    main()
